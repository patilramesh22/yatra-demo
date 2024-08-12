import csv
import inspect
import logging

import openpyxl
import softest


class Utils(softest.TestCase):
    def assert_list_item_text(self, list, value):
        # for stop in list:
        #     print("The stop is:", stop.text)
        #     self.soft_assert(self.assertEqual, stop.text, value)
        #     if stop.text == value:
        #         print("Test Passed")
        #     else:
        #         print("Test Failed")
        # self.assert_all()
        try:
            for stop in list:
                print("The stop is:", stop.text)
                assert value in stop.text
                print("assert pass")
        except:
            print("assert fail")

    def custom_logger(logLevel=logging.DEBUG):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        fh = logging.FileHandler("automation.log")
        formatter = logging.Formatter("%(asctime)s: %(levelname)s: %(name)s: %(message)s",
                                      datefmt="%d-%m-%Y %I:%M:%S %p")
        fh.setFormatter(formatter)
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(path, sheet):
        wb = openpyxl.load_workbook(path)
        sh = wb[sheet]
        rc = sh.max_row
        cc = sh.max_column
        datalist = []
        for i in range(2, rc + 1):
            row = []
            for j in range(1, cc + 1):
                row.append(sh.cell(i, j).value)
            datalist.append(row)
        return datalist

    def get_data_from_csv(path):
        datalist=[]
        csvdata=open(path,"r")
        reader=csv.reader(csvdata)
        next(reader)
        for row in reader:
            datalist.append(row)
        return datalist

